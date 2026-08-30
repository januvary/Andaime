#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generic SQLite database browser/editor (Qt).

Reusable across apps — no business logic, no hardcoded table names.
Auto-detects ``.db`` files in a data directory or accepts explicit paths.

Usage::

    from andaime.qt.dashboard import DashboardService, open_dashboard

    service = DashboardService.from_directory(Path("data"))
    open_dashboard(parent, service, dark_mode=True)
"""

from __future__ import annotations

import sqlite3
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QDialog,
    QFormLayout,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from andaime.qt.theme import get_palette, make_button
from andaime.qt.table import table_batch_populate


def _format_blob_size(length: int | None) -> str:
    """Format a BLOB byte length into a human-readable size string."""
    if length is None or length == 0:
        return "[vazio]"
    if length < 1024:
        return f"{length} B"
    if length < 1024 * 1024:
        return f"{length / 1024:.0f} KB"
    return f"{length / (1024 * 1024):.1f} MB"


@dataclass(frozen=True)
class SearchJoin:
    """Describes a JOIN used only for search/filter on a table.

    Example::

        SearchJoin(
            table="items_catalog",
            on="retirada_items.item_id = items_catalog.item_id",
            search_columns=["items_catalog.descricao"],
        )
    """

    table: str
    on: str
    search_columns: list[str] = field(default_factory=list)


# ======================================================================
# Service
# ======================================================================


class DashboardService:
    """CRUD + introspection layer for browsing SQLite databases."""

    def __init__(
        self,
        database_paths: dict[str, Path | Callable[[], Path]],
        non_editable_columns: dict[str, list[str]] | None = None,
        search_joins: dict[str, list[SearchJoin]] | None = None,
    ) -> None:
        """
        Args:
            database_paths: ``{name: path_or_callable}`` mapping.
                Callable entries are resolved lazily on ``connect_databases``.
            non_editable_columns: Per-table columns that cannot be edited
                (e.g. ``{"pacientes": ["id"]}``).
            search_joins: Per-table JOIN configs for cross-table search
                (e.g. ``{"retirada_items": [SearchJoin(...)]}``).
        """
        self._database_paths = database_paths
        self._non_editable_columns = non_editable_columns or {}
        self._search_joins = search_joins or {}
        self._db_paths: dict[str, Path] = {}

    @classmethod
    def from_directory(
        cls,
        data_dir: Path,
        *,
        non_editable_columns: dict[str, list[str]] | None = None,
        search_joins: dict[str, list[SearchJoin]] | None = None,
    ) -> DashboardService:
        """Auto-detect ``*.db`` files in *data_dir* and build a service.

        The database name is the filename stem (e.g. ``emissor.db`` →
        ``"emissor"``).
        """
        paths: dict[str, Path] = {}
        if data_dir.is_dir():
            for db_file in sorted(data_dir.glob("*.db")):
                paths[db_file.stem] = db_file
        return cls(paths, non_editable_columns, search_joins)

    # ---------- connection ----------

    def connect_databases(self) -> dict[str, Path]:
        """Resolve lazy paths and keep only existing files."""
        self._db_paths = {}
        for name, entry in self._database_paths.items():
            try:
                db_path = entry() if callable(entry) else entry
                if db_path.exists():
                    self._db_paths[name] = db_path
            except Exception:
                pass
        return self._db_paths

    def get_available_databases(self) -> list[str]:
        return list(self._db_paths.keys())

    def _get_connection(self, db_name: str) -> sqlite3.Connection:
        db_path = self._db_paths.get(db_name)
        if not db_path:
            raise ValueError(f"Banco de dados não encontrado: {db_name}")
        conn = sqlite3.connect(str(db_path), timeout=30)
        conn.execute("PRAGMA busy_timeout=30000")
        return conn

    # ---------- introspection ----------

    def _validate_table(self, db_name: str, table_name: str) -> None:
        if table_name not in self._get_db_tables(db_name):
            raise ValueError(
                f"Table '{table_name}' not found in database '{db_name}'"
            )

    def _get_db_tables(self, db_name: str) -> set[str]:
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            return {row[0] for row in cursor.fetchall() if row[0] != "sqlite_sequence"}
        finally:
            conn.close()

    def get_tables(self, db_name: str) -> list[tuple[str, int]]:
        """Return ``(table_name, row_count)`` for every user table."""
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
            result: list[tuple[str, int]] = []
            for row in cursor.fetchall():
                table_name = row[0]
                if table_name == "sqlite_sequence":
                    continue
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                result.append((table_name, int(count)))
            return result
        finally:
            conn.close()

    def _get_table_info(self, db_name: str, table_name: str) -> list[tuple]:
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            return cursor.fetchall()
        finally:
            conn.close()
    def get_table_schema(self, db_name: str, table_name: str) -> dict[str, Any]:
        self._validate_table(db_name, table_name)
        columns_info = self._get_table_info(db_name, table_name)
        column_names = [col[1] for col in columns_info]
        pk_columns = [col[1] for col in columns_info if col[5] > 0]
        blob_columns = {col[1] for col in columns_info if col[2].upper() == "BLOB"}
        return {
            "column_names": column_names,
            "pk_columns": pk_columns,
            "columns_info": columns_info,
            "blob_columns": blob_columns,
        }

    def get_table_rows(
        self, db_name: str, table_name: str, filter_text: str = ""
    ) -> list[tuple]:
        self._validate_table(db_name, table_name)
        conn = self._get_connection(db_name)
        try:
            columns_info = self._get_table_info(db_name, table_name)
            column_names = [col[1] for col in columns_info]
            blob_columns = {col[1] for col in columns_info if col[2].upper() == "BLOB"}

            # Build JOINs for cross-table search
            join_clauses: list[str] = []
            extra_where_parts: list[str] = []
            params: list[Any] = []
            has_joins = False

            if filter_text and table_name in self._search_joins:
                seen_tables: set[str] = set()
                for sj in self._search_joins[table_name]:
                    if sj.table not in seen_tables:
                        join_clauses.append(
                            f"LEFT JOIN {sj.table} ON {sj.on}"
                        )
                        seen_tables.add(sj.table)
                    for col in sj.search_columns:
                        extra_where_parts.append(
                            f"CAST({col} AS TEXT) LIKE ?"
                        )
                        params.append(f"%{filter_text}%")
                has_joins = bool(join_clauses)

            # BLOB columns → LENGTH() (size in bytes, never loads the data)
            # Prefix with table name when JOINs are present to avoid ambiguity
            select_parts = []
            for c in column_names:
                qualified = f"{table_name}.{c}" if has_joins else c
                if c in blob_columns:
                    select_parts.append(f"LENGTH({qualified}) AS {c}")
                else:
                    select_parts.append(
                        f"{qualified} AS {c}" if has_joins else c
                    )

            cursor = conn.cursor()
            cols_sql = ", ".join(select_parts)

            if filter_text:
                # Base WHERE: search all original columns (qualified if JOINs)
                base_where = " OR ".join(
                    f"CAST({table_name}.{col} AS TEXT) LIKE ?"
                    if has_joins
                    else f"CAST({col} AS TEXT) LIKE ?"
                    for col in column_names
                )
                all_where_parts = [base_where]
                params = [f"%{filter_text}%"] * len(column_names) + params
                if extra_where_parts:
                    all_where_parts.append(
                        " OR ".join(extra_where_parts)
                    )
                where = " OR ".join(
                    f"({p})" if " OR " in p else p
                    for p in all_where_parts
                )
                join_sql = f" {' '.join(join_clauses)}" if join_clauses else ""
                cursor.execute(
                    f"SELECT {cols_sql} FROM {table_name}{join_sql} WHERE {where}",
                    params,
                )
            else:
                cursor.execute(f"SELECT {cols_sql} FROM {table_name}")
            return cursor.fetchall()
        finally:
            conn.close()

    def get_non_editable_columns(self, table_name: str) -> list[str]:
        return self._non_editable_columns.get(table_name, [])

    # ---------- mutations ----------

    def update_record(
        self,
        db_name: str,
        table_name: str,
        pk_columns: list[str],
        pk_values: dict[str, Any],
        column_name: str,
        value: Any,
    ) -> None:
        self._validate_table(db_name, table_name)
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE")
            db_value = None if value == "" else value
            where = " AND ".join(f"{pk} = ?" for pk in pk_columns)
            params = [db_value] + [pk_values.get(pk) for pk in pk_columns]
            cursor.execute(
                f"UPDATE {table_name} SET {column_name} = ? WHERE {where}",
                params,
            )
            conn.commit()
        finally:
            conn.close()

    def insert_record(
        self, db_name: str, table_name: str, values: dict[str, Any]
    ) -> None:
        self._validate_table(db_name, table_name)
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE")
            cols = list(values.keys())
            placeholders = ", ".join("?" for _ in cols)
            cursor.execute(
                f"INSERT INTO {table_name} ({', '.join(cols)}) "
                f"VALUES ({placeholders})",
                [values[c] for c in cols],
            )
            conn.commit()
        finally:
            conn.close()

    def delete_record(
        self,
        db_name: str,
        table_name: str,
        pk_columns: list[str],
        pk_values: dict[str, Any],
    ) -> None:
        self._validate_table(db_name, table_name)
        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE")
            where = " AND ".join(f"{pk} = ?" for pk in pk_columns)
            cursor.execute(
                f"DELETE FROM {table_name} WHERE {where}",
                [pk_values.get(pk) for pk in pk_columns],
            )
            conn.commit()
        finally:
            conn.close()

    def parse_integrity_error(
        self,
        error: sqlite3.IntegrityError,
        table: str,
        column: str,
        value: str,
    ) -> str:
        msg = str(error)
        if "UNIQUE constraint failed" in msg:
            return f"Valor '{value}' já existe em '{table}'"
        if "NOT NULL constraint failed" in msg:
            return f"Coluna '{column}' não pode ser vazia"
        if "FOREIGN KEY constraint failed" in msg:
            return f"Referência inválida: '{value}'"
        return f"Erro de banco de dados: {msg}"

    # ---------- export ----------

    def export_to_excel(
        self,
        db_name: str,
        table_name: str,
        save_path: str,
    ) -> str:
        """Export the table to an Excel (.xlsx) file.

        Returns the full path of the saved file.
        """
        self._validate_table(db_name, table_name)

        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl não está instalado") from None

        conn = self._get_connection(db_name)
        try:
            cursor = conn.cursor()

            # Get column names
            columns_info = self._get_table_info(db_name, table_name)
            column_names = [col[1] for col in columns_info]
            blob_columns = {col[1] for col in columns_info if col[2].upper() == "BLOB"}

            # Build SELECT with LENGTH() for blobs
            select_parts = []
            for c in column_names:
                if c in blob_columns:
                    select_parts.append(f"LENGTH({c}) AS {c}")
                else:
                    select_parts.append(c)

            cursor.execute(f"SELECT {', '.join(select_parts)} FROM {table_name}")
            rows = cursor.fetchall()
        finally:
            conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name

        # Header row
        for col_idx, col_name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx, col_name in enumerate(column_names, 1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

        wb.save(save_path)
        return save_path


# ======================================================================
# Window
# ======================================================================


class _AddRecordDialog(QDialog):
    """Minimal insert-record dialog."""

    def __init__(
        self,
        parent: QWidget,
        table_name: str,
        editable_columns: list[tuple],
        palette: dict[str, str],
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(f"Adicionar a {table_name}")
        self.setMinimumWidth(420)
        self._inputs: dict[str, QLineEdit] = {}

        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        form = QFormLayout()
        form.setSpacing(8)
        for col_info in editable_columns:
            col_name = col_info[1]
            entry = QLineEdit()
            entry.setPlaceholderText(col_name)
            form.addRow(f"{col_name}:", entry)
            self._inputs[col_name] = entry
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = make_button("Cancelar", "flat-fill", self)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        ok_btn = make_button("Adicionar", "primary", self)
        ok_btn.clicked.connect(self.accept)
        ok_btn.setDefault(True)
        btn_row.addWidget(ok_btn)
        layout.addLayout(btn_row)

        if self._inputs:
            next(iter(self._inputs.values())).setFocus()

    def values(self) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for col_name, entry in self._inputs.items():
            text = entry.text().strip()
            result[col_name] = text if text else None
        return result


class DashboardWindow(QMainWindow):
    """Generic SQLite browser/editor window."""

    def __init__(
        self,
        parent: QWidget | None,
        service: DashboardService,
        dark_mode: bool = True,
        mask_fn: Callable[[str, str], str] | None = None,
    ) -> None:
        super().__init__(parent)
        self._service = service
        self._mask_fn = mask_fn
        self.setWindowTitle("Painel")
        self.setMinimumSize(1000, 700)
        self.setWindowFlag(Qt.WindowType.Window, True)

        self._palette = get_palette(dark_mode)
        self._db_paths: dict[str, Path] = {}
        self._current_db: str = ""
        self._current_table: str | None = None

        self._column_names: list[str] = []
        self._pk_columns: list[str] = []
        self._pk_values: dict[str, dict[str, Any]] = {}
        self._original_values: dict[str, dict[str, Any]] = {}
        self._unsaved_changes: dict[str, dict[str, Any]] = {}

        self._db_buttons: dict[str, QPushButton] = {}
        self._table_buttons: dict[str, QPushButton] = {}
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(120)
        self._search_timer.timeout.connect(self._execute_search)

        self._setup_ui()
        self._connect_databases()

    # ---------------------------------------------------------------- UI

    def _setup_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_sidebar())
        root.addWidget(self._build_main(), stretch=1)

    def _build_sidebar(self) -> QWidget:
        sidebar = QFrame()
        sidebar.setProperty("class", "panel")
        sidebar.setFixedWidth(240)
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(8, 12, 8, 12)
        layout.setSpacing(8)

        db_label = QLabel("Banco de Dados")
        db_label.setProperty("class", "panel-title")
        layout.addWidget(db_label)

        # DB buttons are created dynamically after connect_databases()
        self._db_buttons_container = QVBoxLayout()
        self._db_buttons_container.setSpacing(4)
        layout.addLayout(self._db_buttons_container)

        tables_label = QLabel("Tabelas")
        tables_label.setProperty("class", "panel-title")
        layout.addWidget(tables_label)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        self._tables_container = QWidget()
        self._tables_layout = QVBoxLayout(self._tables_container)
        self._tables_layout.setContentsMargins(0, 0, 0, 0)
        self._tables_layout.setSpacing(4)
        self._tables_layout.addStretch()
        scroll.setWidget(self._tables_container)
        layout.addWidget(scroll, stretch=1)

        return sidebar

    def _build_main(self) -> QWidget:
        main = QWidget()
        layout = QVBoxLayout(main)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        search_row = QHBoxLayout()
        search_row.setSpacing(8)

        search_row.addWidget(QLabel("Pesquisa:"))

        self._search_entry = QLineEdit()
        self._search_entry.setPlaceholderText("Digite para filtrar...")
        self._search_entry.textChanged.connect(self._on_search)
        search_row.addWidget(self._search_entry, stretch=1)

        self._add_button = make_button("Adicionar", "flat-fill", main)
        self._add_button.setEnabled(False)
        self._add_button.clicked.connect(self._add_record)
        search_row.addWidget(self._add_button)

        self._delete_button = make_button("Excluir", "flat-fill", main)
        self._delete_button.setEnabled(False)
        self._delete_button.clicked.connect(self._delete_record)
        search_row.addWidget(self._delete_button)

        self._save_button = make_button("Salvar (0)", "primary", main)
        self._save_button.clicked.connect(self._save_changes)
        search_row.addWidget(self._save_button)

        self._export_button = make_button("Exportar", "flat-fill", main)
        self._export_button.setEnabled(False)
        self._export_button.clicked.connect(self._export_table)
        search_row.addWidget(self._export_button)

        layout.addLayout(search_row)

        self._table = QTableWidget()
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(True)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self._table.verticalHeader().setVisible(False)
        self._table.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self._table, stretch=1)

        return main

    # -------------------------------------------------------- data / load

    def _connect_databases(self) -> None:
        self._db_paths = self._service.connect_databases()

        # Build sidebar DB buttons dynamically
        for db_name in self._service.get_available_databases():
            btn = make_button(db_name.capitalize(), "flat-fill")
            btn.clicked.connect(
                lambda checked=False, d=db_name: self._switch_database(d)
            )
            self._db_buttons_container.addWidget(btn)
            self._db_buttons[db_name] = btn

        available = self._service.get_available_databases()
        self._current_db = available[0] if available else ""
        self._rebuild_db_buttons()
        if self._current_db:
            self._populate_table_list()

    def _rebuild_db_buttons(self) -> None:
        for db_name, btn in self._db_buttons.items():
            available = db_name in self._db_paths
            btn.setEnabled(available)
            cls = "primary" if db_name == self._current_db and available else "flat-fill"
            btn.setProperty("class", cls)
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def _switch_database(self, db_name: str) -> None:
        if db_name not in self._db_paths:
            QMessageBox.critical(self, "Erro", f"Banco de dados '{db_name}' não disponível.")
            return
        if self._unsaved_changes:
            reply = QMessageBox.question(
                self,
                "Alterações não salvas",
                "Há alterações não salvas. Alternar mesmo assim?",
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        self._current_db = db_name
        self._current_table = None
        self._unsaved_changes.clear()
        self._update_save_button()
        self._rebuild_db_buttons()
        self._table.clear()
        self._table.setRowCount(0)
        self._table.setColumnCount(0)
        self._add_button.setEnabled(False)
        self._delete_button.setEnabled(False)
        self._export_button.setEnabled(False)
        self._populate_table_list()

    def _rebuild_table_buttons(self) -> None:
        for table_name, btn in self._table_buttons.items():
            cls = "primary" if table_name == self._current_table else ""
            btn.setProperty("class", cls)
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def _populate_table_list(self) -> None:
        while self._tables_layout.count() > 1:
            item = self._tables_layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self._table_buttons.clear()

        if self._current_db not in self._db_paths:
            return

        for table_name, count in self._service.get_tables(self._current_db):
            btn = make_button(f"{table_name}\n({count} linhas)", "flat-fill")
            btn.setMinimumHeight(44)
            btn.clicked.connect(
                lambda checked=False, t=table_name: self._select_table(t)
            )
            self._tables_layout.insertWidget(self._tables_layout.count() - 1, btn)
            self._table_buttons[table_name] = btn

        self._rebuild_table_buttons()

    def _select_table(self, table_name: str) -> None:
        if self._current_table == table_name:
            return
        if self._unsaved_changes:
            reply = QMessageBox.question(
                self,
                "Alterações não salvas",
                "Há alterações não salvas. Alternar mesmo assim?",
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        self._current_table = table_name
        self._rebuild_table_buttons()
        self._search_entry.clear()
        self._add_button.setEnabled(True)
        self._delete_button.setEnabled(True)
        self._export_button.setEnabled(True)
        self._unsaved_changes.clear()
        self._update_save_button()
        self._populate_table(table_name)

    def _populate_table(self, table_name: str, filter_text: str = "") -> None:
        if self._current_db not in self._db_paths:
            return

        try:
            schema = self._service.get_table_schema(self._current_db, table_name)
            rows = self._service.get_table_rows(
                self._current_db, table_name, filter_text
            )
        except (ValueError, sqlite3.Error) as e:
            QMessageBox.critical(self, "Erro", str(e))
            self._table.clear()
            self._table.setRowCount(0)
            self._table.setColumnCount(0)
            return

        blob_columns = schema["blob_columns"]
        self._column_names = schema["column_names"]
        self._pk_columns = schema["pk_columns"]
        non_editable = set(self._service.get_non_editable_columns(table_name)) | set(self._pk_columns)

        with table_batch_populate(self._table):
            self._table.clear()
            self._table.setColumnCount(len(self._column_names))
            self._table.setHorizontalHeaderLabels(self._column_names)
            self._table.setRowCount(len(rows))

            self._pk_values.clear()
            self._original_values.clear()

            for row_idx, row in enumerate(rows):
                row_key = str(row_idx)
                pk_vals: dict[str, Any] = {}
                originals: dict[str, Any] = {}

                for col_idx, col_name in enumerate(self._column_names):
                    raw = row[col_idx]
                    originals[col_name] = raw
                    if col_name in self._pk_columns:
                        pk_vals[col_name] = raw

                    if col_name in blob_columns:
                        display = _format_blob_size(raw)
                    elif raw is None:
                        display = "-"
                    else:
                        display = str(raw)
                        if col_name not in non_editable and self._mask_fn is not None:
                            masked = self._mask_fn(col_name, str(raw))
                            if masked:
                                display = masked

                    item = QTableWidgetItem(display)
                    item.setData(Qt.ItemDataRole.UserRole, row_key)
                    if col_name in non_editable or col_name in blob_columns:
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        item.setForeground(QColor(self._palette["text_dim"]))
                    self._table.setItem(row_idx, col_idx, item)

                self._pk_values[row_key] = pk_vals
                self._original_values[row_key] = originals

        self._table.resizeColumnsToContents()

    # ------------------------------------------------------------- search

    def _on_search(self, _text: str = "") -> None:
        self._search_timer.start()

    def _execute_search(self) -> None:
        if self._current_table:
            self._populate_table(self._current_table, self._search_entry.text())

    # ----------------------------------------------------------- editing

    def _row_key_for_item(self, item: QTableWidgetItem) -> str | None:
        key = item.data(Qt.ItemDataRole.UserRole)
        return str(key) if key is not None else None

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        """Track inline edits as unsaved changes.

        Convention: ``None`` = empty (db), ``"-"`` = display-empty (table).
        Typing ``"-"`` or ``""`` into an empty cell = no change.
        """
        if self._current_table is None:
            return

        col = item.column()
        if col < 0 or col >= len(self._column_names):
            return

        col_name = self._column_names[col]
        non_editable = self._service.get_non_editable_columns(self._current_table)
        if col_name in non_editable:
            return

        row_key = self._row_key_for_item(item)
        if row_key is None:
            return

        new_text = item.text()
        new_raw = None if new_text in ("", "-") else new_text

        original = self._original_values.get(row_key, {}).get(col_name)

        if new_raw == original:
            if row_key in self._unsaved_changes:
                self._unsaved_changes[row_key].pop(col_name, None)
                if not self._unsaved_changes[row_key]:
                    del self._unsaved_changes[row_key]
            self._update_save_button()
            return

        if new_raw is not None and self._mask_fn is not None:
            masked = self._mask_fn(col_name, new_raw) or new_raw
            if masked != item.text():
                self._table.blockSignals(True)
                item.setText(masked)
                self._table.blockSignals(False)

        self._unsaved_changes.setdefault(row_key, {})[col_name] = (
            new_raw if new_raw is not None else ""
        )
        self._update_save_button()

    def _update_save_button(self) -> None:
        total = sum(len(c) for c in self._unsaved_changes.values())
        self._save_button.setText(f"Salvar ({total})")

    def _save_changes(self) -> None:
        if not self._unsaved_changes or not self._current_table:
            return

        saved = 0
        changes_copy = {
            rk: dict(cols) for rk, cols in self._unsaved_changes.items()
        }

        for row_key, changes in changes_copy.items():
            pk_vals = self._pk_values.get(row_key, {})
            for col_name, new_value in list(changes.items()):
                value = new_value
                try:
                    self._service.update_record(
                        self._current_db,
                        self._current_table,
                        self._pk_columns or ["id"],
                        pk_vals,
                        col_name,
                        value,
                    )
                    saved += 1
                    self._original_values.setdefault(row_key, {})[col_name] = (
                        new_value if new_value != "" else None
                    )
                    del self._unsaved_changes[row_key][col_name]
                    if not self._unsaved_changes[row_key]:
                        del self._unsaved_changes[row_key]
                except sqlite3.IntegrityError as e:
                    original = self._original_values.get(row_key, {}).get(col_name)
                    display = "-" if original is None else str(original)
                    col_idx = self._column_names.index(col_name)
                    for r in range(self._table.rowCount()):
                        it = self._table.item(r, col_idx)
                        if it and self._row_key_for_item(it) == row_key:
                            self._table.blockSignals(True)
                            it.setText(display)
                            self._table.blockSignals(False)
                            break
                    msg = self._service.parse_integrity_error(
                        e, self._current_table, col_name, str(new_value)
                    )
                    QMessageBox.critical(self, "Erro ao salvar", msg)
                    del self._unsaved_changes[row_key][col_name]
                    if not self._unsaved_changes[row_key]:
                        del self._unsaved_changes[row_key]

        self._update_save_button()
        if saved > 0:
            QMessageBox.information(
                self, "Salvo", f"{saved} alteração(ões) salva(s)."
            )

    # ------------------------------------------------------------- export

    def _export_table(self) -> None:
        if not self._current_table:
            return

        from datetime import datetime

        default_name = f"{self._current_table}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar para Excel",
            str(Path.home() / "Downloads" / default_name),
            "Planilha Excel (*.xlsx)",
        )
        if not save_path:
            return

        try:
            self._service.export_to_excel(
                self._current_db, self._current_table, save_path
            )
            QMessageBox.information(
                self, "Exportado", f"Planilha salva em:\n{save_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro ao exportar", str(e))

    # -------------------------------------------------------------- CRUD

    def _add_record(self) -> None:
        if not self._current_table:
            return

        schema = self._service.get_table_schema(self._current_db, self._current_table)
        columns_info = schema["columns_info"]
        non_editable = set(
            self._service.get_non_editable_columns(self._current_table)
        )
        editable = [
            col
            for col in columns_info
            if col[1] not in non_editable
            and not (col[5] > 0 and str(col[2]).upper() == "INTEGER")
        ]
        if not editable:
            QMessageBox.warning(
                self,
                "Sem colunas editáveis",
                f"Tabela '{self._current_table}' não possui colunas editáveis.",
            )
            return

        dialog = _AddRecordDialog(self, self._current_table, editable, self._palette)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        values = dialog.values()

        try:
            self._service.insert_record(self._current_db, self._current_table, values)
            self._populate_table(self._current_table)
            self._populate_table_list()
            QMessageBox.information(self, "Adicionado", "Registro adicionado com sucesso.")
        except sqlite3.IntegrityError as e:
            msg = self._service.parse_integrity_error(e, self._current_table, "", "")
            QMessageBox.critical(self, "Erro ao adicionar", msg)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro inesperado: {e}")

    def _delete_record(self) -> None:
        if not self._current_table:
            return

        selected = self._table.selectedItems()
        if not selected:
            QMessageBox.warning(
                self, "Nenhuma seleção", "Selecione um registro para excluir."
            )
            return

        row = selected[0].row()
        first_item = self._table.item(row, 0)
        if first_item is None:
            return
        row_key = self._row_key_for_item(first_item)
        if row_key is None:
            return

        pk_vals = self._pk_values.get(row_key, {})
        pk_cols = self._pk_columns or ["id"]
        pk_str = ", ".join(f"{k}={v}" for k, v in pk_vals.items() if k in pk_cols)

        reply = QMessageBox.question(
            self,
            "Confirmar exclusão",
            f"Excluir registro:\n{self._current_table}[{pk_str}]?",
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            self._service.delete_record(
                self._current_db, self._current_table, pk_cols, pk_vals
            )
            self._unsaved_changes.pop(row_key, None)
            self._update_save_button()
            self._populate_table(self._current_table)
            self._populate_table_list()
            QMessageBox.information(self, "Excluído", "Registro excluído com sucesso.")
        except sqlite3.IntegrityError as e:
            msg = self._service.parse_integrity_error(e, self._current_table, "", "")
            QMessageBox.critical(self, "Erro ao excluir", msg)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro inesperado: {e}")


# ======================================================================
# Convenience
# ======================================================================


def open_dashboard(
    parent: QWidget,
    service: DashboardService,
    dark_mode: bool = True,
    mask_fn: Callable[[str, str], str] | None = None,
) -> DashboardWindow:
    """Create and show a :class:`DashboardWindow`.

    Args:
        parent: Parent widget.
        service: Configured :class:`DashboardService` instance.
        dark_mode: Use dark theme.
        mask_fn: Optional ``(column_name, raw_value) -> display_value``.
    """
    window = DashboardWindow(parent, service, dark_mode, mask_fn)
    window.show()
    return window
