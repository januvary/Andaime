"""Exporta os processos do banco principal (ss54.db) para uma planilha Excel.

Uso:
    PYTHONPATH=. python -m src.utils.export_to_xlsx

O formato segue o layout da planilha original: uma aba por remessa, com
seções "PRIMEIRA SOLICITAÇÃO" e "RENOVAÇÃO" e as colunas clássicas do
processo administrativo SS-54.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from andaime.dates import parse_date
from bap.constants import SOLICITACAO_LABELS, TIPO_UPPER, STATUS_LABELS
from bap.database.ss54_database import SS54Database
from bap.utils.text_utils import format_phone
from bap.utils.date_utils import format_date_display

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CALIBRI = Font(name="Calibri")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


SECTION_FILLS = {
    "primeira": _fill("71A8DA"),
    "renovacao": _fill("FFCC33"),
}

STATUS_COLORS = {
    "em_analise": "F2F2F2",
    "incompleto": "FCE4D6",
    "completo": "C6EFCE",
    "enviado": "BDD7EE",
    "correcao": "FFEB9C",
    "autorizado": "A9D08E",
    "expirado": "D9D9D9",
    "negado": "FFC7CE",
    "encerrado": "BFBFBF",
}

DEFAULT_OUTPUT = Path("data") / "processos_export.xlsx"

COLUMNS = [
    "NOME PACIENTE",
    "TIPO",
    "DESCRIÇÃO",
    "RETORNO DRS-IV",
    "TELEFONE",
    "STATUS",
]


def _sheet_name_from_date(date_str: str) -> str:
    """Retorna nome de aba no estilo da planilha original (dd-mm-aaaa)."""
    dt = parse_date(date_str)
    if not dt:
        return date_str[:31]
    return dt.strftime("%d-%m-%Y")


def _format_tipo(tipo: str | None) -> str:
    """Exibe tipo em maiúsculas, próximo ao estilo da planilha original."""
    return TIPO_UPPER.get(tipo, (tipo or "").upper())


def _set_column_widths(ws) -> None:
    widths = {
        "A": 35,  # NOME PACIENTE
        "B": 20,  # TIPO
        "C": 35,  # DESCRIÇÃO
        "D": 60,  # RETORNO DRS-IV
        "E": 18,  # TELEFONE
        "F": 18,  # STATUS
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def export_processos_to_xlsx(
    db: SS54Database,
    output_path: str | Path = DEFAULT_OUTPUT,
) -> str:
    """Exporta processos para uma planilha Excel com uma aba por remessa."""
    rows = db.get_processos_for_export()

    # Agrupa por remessa e, dentro dela, por tipo de solicitação
    grouped: dict[str, dict[str, list[dict]]] = {}
    for r in rows:
        remessa = r["lote_date"]
        solicitacao = r["solicitacao"] or "primeira"
        grouped.setdefault(remessa, {}).setdefault(solicitacao, []).append(r)

    wb = openpyxl.Workbook()
    # Remove a aba padrão criada automaticamente; criaremos uma aba por remessa
    default_sheet = wb.active
    wb.remove(default_sheet)

    for remessa in sorted(grouped.keys()):
        sheet_name = _sheet_name_from_date(remessa)
        ws = wb.create_sheet(title=sheet_name)
        _set_column_widths(ws)

        # Título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "PROCESSO ADMNISTRATIVO SS54"
        title_cell.font = Font(name="Calibri", bold=True, size=14)
        title_cell.alignment = CENTER

        # Remessa
        ws["A3"] = "REMESSA:"
        ws["B3"] = format_date_display(remessa)
        ws["A3"].font = Font(name="Calibri", bold=True)
        ws["A3"].alignment = CENTER
        ws["B3"].alignment = CENTER

        current_row = 5
        for solicitacao in ["primeira", "renovacao"]:
            section_rows = grouped[remessa].get(solicitacao, [])
            if not section_rows:
                continue

            # Cabeçalho da seção (mesclado com borda em todos os lados + cor)
            fill = SECTION_FILLS.get(solicitacao)
            for col in range(1, 7):
                c = ws.cell(row=current_row, column=col)
                c.font = Font(name="Calibri", bold=True, size=16, color="000000")
                c.alignment = CENTER
                c.border = BORDER
                if fill is not None:
                    c.fill = fill
            ws.cell(row=current_row, column=1).value = SOLICITACAO_LABELS.get(
                solicitacao, solicitacao.upper()
            ).upper()
            ws.merge_cells(
                start_row=current_row, start_column=1, end_row=current_row, end_column=6
            )
            current_row += 1

            # Cabeçalhos das colunas
            for col_idx, header in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(name="Calibri", bold=True)
                cell.alignment = CENTER
                cell.border = BORDER
            current_row += 1

            # Dados
            for r in section_rows:
                raw_status = r["status"] or ""
                values = [
                    r["paciente_nome"] or "",
                    _format_tipo(r["tipo"]),
                    r["descricao"] or "",
                    r["observacoes"] or "",
                    format_phone(r["paciente_telefone"]) or "",
                    STATUS_LABELS.get(raw_status, raw_status or ""),
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.font = CALIBRI
                    cell.alignment = LEFT if col_idx in (1, 4) else CENTER
                    cell.border = BORDER
                    if col_idx == 6:
                        color = STATUS_COLORS.get(raw_status)
                        if color:
                            cell.fill = _fill(color)
                current_row += 1

            current_row += 1  # linha em branco entre seções

    wb.save(output_path)
    return str(output_path)


def main() -> None:
    from bap.utils.bootstrap import ensure_initialized

    ensure_initialized()

    db = SS54Database()
    try:
        path = export_processos_to_xlsx(db)
        print(f"Exportado: {path}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
